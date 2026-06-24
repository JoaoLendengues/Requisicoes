"""Cadastro de clientes: individual + importação em lote por planilha Excel.

Embarcado como aba "Clientes" nas Configurações (admin). Diferente da Central
de Usuários, a lista é orientada por busca no servidor — a base tem dezenas de
milhares de clientes e não pode ser carregada inteira no app.
"""

from PySide6.QtCore import QObject, QThread, Qt, QTimer, Signal
from PySide6.QtGui import QColor, QPalette
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..api import client as api
from ..core import theme
from ..core.dialogs import ask_confirmation
from ..core.datetime_utils import (
    format_datetime as _format_datetime,
    format_header_date as _format_header_date,
    local_now,
)
from ..core.text_case import bind_uppercase_line_edit
from ..widgets.smooth_scroll import SmoothScrollArea, apply_smooth_scroll

# Reaproveita os estilos/helpers já usados na Central de Usuários para manter
# a aparência consistente entre as abas de Configurações.
from .user_center_view import (
    ActionWorker,
    UiCallback,
    _danger_action_btn_style,
    _field_style,
    _flat_secondary_btn_style,
    _make_card,
    _primary_action_btn_style,
    _rgba,
)

SEARCH_LIMIT = 100
_HEADER_ALIASES = {
    "code": "code", "codigo": "code", "código": "code",
    "name": "name", "nome": "name",
    "cnpj": "cnpj", "cpf_cnpj": "cnpj", "cpf/cnpj": "cnpj", "cpf": "cnpj",
}


class ImportWorker(QObject):
    """Cria os clientes da planilha um a um (modo create-only).

    Cada linha vira um POST /clients/. Sucesso conta como criado; conflito
    (código/CNPJ já existente) é registrado como rejeitado, sem abortar o resto.
    """

    progress = Signal(int, int)      # (concluídas, total)
    finished = Signal(dict)

    def __init__(self, rows: list[dict]):
        super().__init__()
        self.rows = rows
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def run(self):
        created = 0
        rejected: list[tuple[str, str]] = []
        total = len(self.rows)
        for i, row in enumerate(self.rows, start=1):
            if self._cancel:
                break
            try:
                api.create_client(row)
                created += 1
            except api.APIError as exc:
                rejected.append((row.get("code", ""), exc.detail))
            except Exception as exc:  # noqa: BLE001
                rejected.append((row.get("code", ""), str(exc)))
            self.progress.emit(i, total)
        self.finished.emit(
            {"created": created, "rejected": rejected, "cancelled": self._cancel}
        )


class ClientCenterView(QWidget):
    guide_requested = Signal()

    def __init__(self, scale: float = 1.0, parent=None, embedded: bool = True):
        super().__init__(parent)
        self.scale = scale
        self.embedded = embedded
        self._threads: list[tuple[QThread, QObject]] = []
        self._clients_visible: list[dict] = []
        self._selected_client_id: int | None = None
        self._import_thread: QThread | None = None
        self._import_worker: ImportWorker | None = None
        self._progress: QProgressDialog | None = None

        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(320)
        self._search_timer.timeout.connect(self._do_search)

        self._setup_ui()

    # ── Construção da UI ──────────────────────────────────────────────────────
    def _setup_ui(self):
        s = self.scale
        page_bg = theme.CONTENT_BG
        self.setObjectName("clientCenterView")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(f"QWidget#clientCenterView {{ background:{page_bg}; }}")

        root = QVBoxLayout(self)
        if self.embedded:
            root.setContentsMargins(0, 0, 0, 0)
        else:
            m = max(18, int(24 * s))
            root.setContentsMargins(m, m, m, m)
        root.setSpacing(max(14, int(18 * s)))

        # No modo embarcado, Configurações já provê título/data; mantemos os
        # widgets de cabeçalho como ocultos para a lógica de status continuar.
        self.date_label = QLabel(_format_header_date())
        self.updated_label = QLabel("")
        self.refresh_btn = QPushButton()
        self.refresh_btn.hide()

        self.error_label = QLabel("")
        self.error_label.hide()
        self.error_label.setWordWrap(True)
        self.error_label.setStyleSheet(
            f"background:{_rgba(theme.DANGER, 18)}; color:{theme.DANGER};"
            f"border:1px solid {_rgba(theme.DANGER, 48)}; border-radius:16px;"
            f"padding:12px 14px; font-size:{max(8, int(9 * s))}pt; font-weight:600;"
        )
        root.addWidget(self.error_label)

        self._page_scroll = SmoothScrollArea()
        self._page_scroll.setWidgetResizable(True)
        self._page_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._page_scroll.setStyleSheet(f"QScrollArea {{ border:none; background:{page_bg}; }}")
        self._page_scroll.viewport().setStyleSheet(f"background:{page_bg}; border:none;")
        root.addWidget(self._page_scroll, 1)

        self._page_content = QWidget()
        self._page_content.setObjectName("clientCenterContent")
        self._page_content.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self._page_content.setStyleSheet(
            f"QWidget#clientCenterContent {{ background:{page_bg}; }}"
        )
        self._page_scroll.setWidget(self._page_content)

        layout = QVBoxLayout(self._page_content)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(max(16, int(18 * s)))

        body = QHBoxLayout()
        body.setSpacing(max(12, int(14 * s)))
        body.addWidget(self._build_table_card(), 3)
        body.addWidget(self._build_form_card(), 2)
        layout.addLayout(body)
        layout.addStretch()

        self._prepare_new_client()  # estado inicial coerente do formulário

    def _build_table_card(self) -> QFrame:
        s = self.scale
        card = _make_card(s, theme.CARD_BG, radius=max(18, int(20 * s)))
        layout = QVBoxLayout(card)
        layout.setContentsMargins(max(16, int(20 * s)), max(14, int(18 * s)),
                                  max(16, int(20 * s)), max(14, int(18 * s)))
        layout.setSpacing(max(10, int(12 * s)))

        header = QHBoxLayout()
        header.setSpacing(max(8, int(10 * s)))
        title = QLabel("CLIENTES CADASTRADOS")
        title.setStyleSheet(
            f"font-size:{max(10, int(12 * s))}pt; font-weight:800; background:transparent;"
        )
        header.addWidget(title)
        header.addStretch()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar por nome, código ou CNPJ...")
        self.search_input.setFixedHeight(max(38, int(44 * s)))
        self.search_input.setStyleSheet(_field_style(s))
        self.search_input.textChanged.connect(lambda _: self._search_timer.start())
        header.addWidget(self.search_input, 1)

        new_btn = QPushButton("NOVO")
        new_btn.setFixedHeight(max(38, int(44 * s)))
        new_btn.setStyleSheet(_flat_secondary_btn_style(s))
        new_btn.clicked.connect(self._prepare_new_client)
        header.addWidget(new_btn)
        layout.addLayout(header)

        # Linha de ações de importação
        import_row = QHBoxLayout()
        import_row.setSpacing(max(8, int(10 * s)))
        self.btn_template = QPushButton("BAIXAR MODELO")
        self.btn_template.setFixedHeight(max(34, int(40 * s)))
        self.btn_template.setStyleSheet(_flat_secondary_btn_style(s))
        self.btn_template.clicked.connect(self._download_template)
        self.btn_import = QPushButton("IMPORTAR PLANILHA")
        self.btn_import.setFixedHeight(max(34, int(40 * s)))
        self.btn_import.setStyleSheet(_flat_secondary_btn_style(s))
        self.btn_import.clicked.connect(self._import_excel)
        import_row.addWidget(self.btn_template)
        import_row.addWidget(self.btn_import)
        import_row.addStretch()
        layout.addLayout(import_row)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["CÓDIGO", "NOME", "CNPJ", "STATUS"])
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setFrameShape(QFrame.Shape.NoFrame)
        self.table.setShowGrid(False)
        apply_smooth_scroll(self.table)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.doubleClicked.connect(self._load_selected_client)
        self.table.itemSelectionChanged.connect(self._load_current_selection)
        head = self.table.horizontalHeader()
        head.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        head.setStretchLastSection(True)
        self.table.setColumnWidth(1, max(200, int(220 * s)))
        head.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        head.setMinimumHeight(max(34, int(40 * s)))
        self.table.verticalHeader().setDefaultSectionSize(max(32, int(38 * s)))
        self._apply_table_style()
        layout.addWidget(self.table, 1)

        self.result_hint = QLabel("")
        self.result_hint.setProperty("muted", "1")
        self.result_hint.setStyleSheet(f"background:transparent; font-size:{max(7, int(8 * s))}pt;")
        layout.addWidget(self.result_hint)
        return card

    def _build_form_card(self) -> QFrame:
        s = self.scale
        card = _make_card(s, theme.CARD_BG, radius=max(18, int(20 * s)))
        layout = QVBoxLayout(card)
        layout.setContentsMargins(max(16, int(20 * s)), max(14, int(18 * s)),
                                  max(16, int(20 * s)), max(14, int(18 * s)))
        layout.setSpacing(max(10, int(12 * s)))

        title = QLabel("CADASTRO INDIVIDUAL")
        title.setStyleSheet(
            f"font-size:{max(10, int(12 * s))}pt; font-weight:800; background:transparent;"
        )
        helper = QLabel(
            "Código, Nome e CNPJ são obrigatórios. O código não pode ser alterado "
            "depois de criado."
        )
        helper.setWordWrap(True)
        helper.setProperty("muted", "1")
        helper.setStyleSheet(f"background:transparent; font-size:{max(7, int(8 * s))}pt;")
        layout.addWidget(title)
        layout.addWidget(helper)

        self.form_status = QLabel("Novo cliente")
        self.form_status.setProperty("accent", "1")
        self.form_status.setStyleSheet(f"background:transparent; font-size:{max(8, int(9 * s))}pt; font-weight:700;")
        layout.addWidget(self.form_status)

        grid = QGridLayout()
        grid.setHorizontalSpacing(max(8, int(10 * s)))
        grid.setVerticalSpacing(max(8, int(10 * s)))

        self.input_code = self._input()
        self.input_name = self._input()
        self.input_cnpj = self._input()
        self.input_cnpj.setPlaceholderText("00.000.000/0000-00 ou CPF")
        bind_uppercase_line_edit(self.input_code)
        bind_uppercase_line_edit(self.input_name)

        grid.addWidget(self._field_label("Código"), 0, 0)
        grid.addWidget(self.input_code, 0, 1)
        grid.addWidget(self._field_label("Nome"), 1, 0)
        grid.addWidget(self.input_name, 1, 1)
        grid.addWidget(self._field_label("CNPJ / CPF"), 2, 0)
        grid.addWidget(self.input_cnpj, 2, 1)
        layout.addLayout(grid)

        self.check_active = QCheckBox("Cliente ativo")
        self.check_active.setChecked(True)
        self.check_active.setStyleSheet(f"background:transparent; font-size:{max(8, int(9 * s))}pt;")
        layout.addWidget(self.check_active)

        actions = QHBoxLayout()
        self.btn_save = QPushButton("SALVAR")
        self.btn_save.setFixedHeight(max(38, int(44 * s)))
        self.btn_save.setStyleSheet(_primary_action_btn_style(s))
        self.btn_save.clicked.connect(self._save_client)
        actions.addWidget(self.btn_save)

        self.btn_disable = QPushButton("DESATIVAR")
        self.btn_disable.setFixedHeight(max(38, int(44 * s)))
        self.btn_disable.setStyleSheet(_danger_action_btn_style(s))
        self.btn_disable.clicked.connect(self._deactivate_client)
        self.btn_disable.setEnabled(False)
        actions.addWidget(self.btn_disable)

        clear_btn = QPushButton("LIMPAR")
        clear_btn.setFixedHeight(max(38, int(44 * s)))
        clear_btn.setStyleSheet(_flat_secondary_btn_style(s))
        clear_btn.clicked.connect(self._prepare_new_client)
        actions.addWidget(clear_btn)
        layout.addLayout(actions)
        layout.addStretch()
        return card

    def _input(self) -> QLineEdit:
        s = self.scale
        field = QLineEdit()
        field.setFixedHeight(max(38, int(44 * s)))
        field.setStyleSheet(_field_style(s))
        return field

    def _field_label(self, text: str) -> QLabel:
        label = QLabel(text.upper())
        label.setProperty("muted", "1")
        label.setStyleSheet(
            f"font-size:{max(7, int(8 * self.scale))}pt; font-weight:700;"
            f"background:transparent;"
        )
        return label

    # ── Busca / listagem ──────────────────────────────────────────────────────
    def refresh(self):
        """Recarrega a lista mantendo o termo de busca atual."""
        self._do_search()

    def _do_search(self):
        term = self.search_input.text().strip()
        self.error_label.hide()
        self._run_action(
            api.list_clients, term, SEARCH_LIMIT, on_result=self._populate_clients
        )

    def _populate_clients(self, payload: object):
        self._clients_visible = payload if isinstance(payload, list) else []
        self._fill_table()
        n = len(self._clients_visible)
        if n >= SEARCH_LIMIT:
            self.result_hint.setText(
                f"Mostrando os primeiros {SEARCH_LIMIT} resultados — refine a busca."
            )
        elif self.search_input.text().strip():
            self.result_hint.setText(f"{n} resultado(s).")
        else:
            self.result_hint.setText(f"{n} cliente(s). Use a busca para localizar outros.")
        current = local_now()
        self.date_label.setText(_format_header_date(current))
        self.updated_label.setText(f"Atualizado em {_format_datetime(current)}")

    def _fill_table(self):
        self.table.setRowCount(0)
        for cli in self._clients_visible:
            if not isinstance(cli, dict):
                continue
            row = self.table.rowCount()
            self.table.insertRow(row)
            values = [
                str(cli.get("code") or "-"),
                str(cli.get("name") or "-"),
                str(cli.get("cnpj") or "-"),
                "Ativo" if cli.get("is_active") else "Inativo",
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, col, item)

    # ── Seleção / formulário ──────────────────────────────────────────────────
    def _prepare_new_client(self):
        self._selected_client_id = None
        self.form_status.setText("Novo cliente")
        self.input_code.clear()
        self.input_name.clear()
        self.input_cnpj.clear()
        self.input_code.setReadOnly(False)
        self.input_code.setStyleSheet(_field_style(self.scale))
        self.check_active.setChecked(True)
        self.check_active.setEnabled(False)  # só relevante ao editar
        self.btn_disable.setEnabled(False)

    def _load_selected_client(self, index):
        row = index.row()
        if 0 <= row < len(self._clients_visible):
            self._load_client_into_form(self._clients_visible[row])

    def _load_current_selection(self):
        row = self.table.currentRow()
        if 0 <= row < len(self._clients_visible):
            self._load_client_into_form(self._clients_visible[row])

    def _load_client_into_form(self, cli: dict):
        self._selected_client_id = int(cli["id"])
        self.form_status.setText("Cadastro carregado")
        self.input_code.setText(str(cli.get("code") or ""))
        self.input_name.setText(str(cli.get("name") or ""))
        self.input_cnpj.setText(str(cli.get("cnpj") or ""))
        # Código é imutável após criado (PATCH não altera código).
        self.input_code.setReadOnly(True)
        self.input_code.setStyleSheet(
            _field_style(self.scale) + "QLineEdit { color:" + theme.TEXT_MEDIUM + "; }"
        )
        self.check_active.setChecked(bool(cli.get("is_active")))
        self.check_active.setEnabled(True)
        self.btn_disable.setEnabled(bool(cli.get("is_active")))

    def _save_client(self):
        code = self.input_code.text().strip()
        name = self.input_name.text().strip()
        cnpj = self.input_cnpj.text().strip()

        if not code or not name or not cnpj:
            QMessageBox.warning(
                self, "Cadastro de clientes",
                "Código, Nome e CNPJ são obrigatórios.",
            )
            return

        if self._selected_client_id is None:
            payload = {"code": code, "name": name, "cnpj": cnpj}
            self._run_action(
                api.create_client, payload,
                on_result=lambda _: self._after_save(code),
                success_message="Cliente cadastrado com sucesso.",
            )
        else:
            payload = {
                "name": name,
                "cnpj": cnpj,
                "is_active": self.check_active.isChecked(),
            }
            self._run_action(
                api.update_client, self._selected_client_id, payload,
                on_result=lambda _: self._after_save(code),
                success_message="Cadastro atualizado com sucesso.",
            )

    def _after_save(self, code: str):
        self.search_input.setText(code)
        self._do_search()
        self._prepare_new_client()

    def _deactivate_client(self):
        if self._selected_client_id is None:
            QMessageBox.information(self, "Cadastro de clientes", "Selecione um cliente primeiro.")
            return
        if not ask_confirmation(
            self, "Cadastro de clientes",
            "Deseja desativar este cliente?", yes_text="Sim", no_text="Não",
        ):
            return
        self._run_action(
            api.deactivate_client, self._selected_client_id,
            on_result=lambda _: (self._do_search(), self._prepare_new_client()),
            success_message="Cliente desativado com sucesso.",
        )

    # ── Importação por planilha ───────────────────────────────────────────────
    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar modelo", "modelo_clientes.xlsx", "Planilha Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "clientes"
            ws.append(["code", "name", "cnpj"])
            ws.append(["EX001", "CLIENTE EXEMPLO LTDA", "00.000.000/0001-00"])
            wb.save(path)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "Baixar modelo", f"Não foi possível salvar o modelo.\n\n{exc}")
            return
        QMessageBox.information(
            self, "Baixar modelo",
            "Modelo salvo com as colunas: code, name, cnpj.\n"
            "A linha de exemplo pode ser removida antes de importar.",
        )

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Escolher planilha", "", "Planilha Excel (*.xlsx *.xlsm)"
        )
        if not path:
            return
        try:
            valid, rejected = self._parse_excel(path)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "Importar planilha", f"Não foi possível ler a planilha.\n\n{exc}")
            return

        if not valid and not rejected:
            QMessageBox.information(self, "Importar planilha", "Nenhuma linha encontrada na planilha.")
            return

        msg = f"{len(valid)} cliente(s) prontos para importar."
        if rejected:
            msg += (f"\n{len(rejected)} linha(s) serão rejeitadas antes do envio "
                    "(duplicadas na planilha ou sem os 3 campos).")
        msg += "\n\nDeseja continuar?"
        if not ask_confirmation(self, "Importar planilha", msg, yes_text="Importar", no_text="Cancelar"):
            return

        if not valid:
            self._show_import_summary({"created": 0, "rejected": [], "cancelled": False}, rejected)
            return

        self._run_import(valid, rejected)

    def _parse_excel(self, path: str):
        """Lê a planilha e devolve (linhas_válidas, linhas_pré-rejeitadas)."""
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                raise ValueError("A planilha está vazia.")

            col: dict[str, int] = {}
            for idx, h in enumerate(header):
                key = str(h or "").strip().lower()
                mapped = _HEADER_ALIASES.get(key)
                if mapped and mapped not in col:
                    col[mapped] = idx

            missing = [c for c in ("code", "name", "cnpj") if c not in col]
            if missing:
                raise ValueError(
                    "A planilha precisa das colunas: code, name, cnpj.\n"
                    f"Faltando: {', '.join(missing)}"
                )

            valid: list[dict] = []
            rejected: list[tuple[str, str]] = []
            seen_code: set[str] = set()
            seen_cnpj: set[str] = set()

            def _cell(r, key):
                i = col[key]
                if i < len(r) and r[i] is not None:
                    return str(r[i]).strip()
                return ""

            for rownum, r in enumerate(rows_iter, start=2):
                if r is None:
                    continue
                code = _cell(r, "code").upper()
                name = _cell(r, "name").upper()
                cnpj = _cell(r, "cnpj")
                if not code and not name and not cnpj:
                    continue  # linha em branco
                if not (code and name and cnpj):
                    rejected.append((code or f"linha {rownum}",
                                     "campos obrigatórios faltando (code/name/cnpj)"))
                    continue
                cdig = "".join(ch for ch in cnpj if ch.isdigit())
                if code in seen_code:
                    rejected.append((code, "código duplicado na planilha"))
                    continue
                if cdig and cdig in seen_cnpj:
                    rejected.append((code, "CNPJ duplicado na planilha"))
                    continue
                seen_code.add(code)
                if cdig:
                    seen_cnpj.add(cdig)
                valid.append({"code": code, "name": name, "cnpj": cnpj})
            return valid, rejected
        finally:
            wb.close()

    def _run_import(self, rows: list[dict], pre_rejected: list[tuple[str, str]]):
        total = len(rows)
        self._progress = QProgressDialog("Importando clientes...", "Cancelar", 0, total, self)
        self._progress.setWindowTitle("Importar planilha")
        self._progress.setWindowModality(Qt.WindowModality.WindowModal)
        self._progress.setMinimumDuration(0)
        self._progress.setValue(0)

        self._import_thread = QThread()
        self._import_worker = ImportWorker(rows)
        self._import_worker.moveToThread(self._import_thread)
        self._import_thread.started.connect(self._import_worker.run)
        self._import_worker.progress.connect(self._on_import_progress)
        self._import_worker.finished.connect(
            lambda result: self._on_import_finished(result, pre_rejected)
        )
        self._import_worker.finished.connect(self._import_thread.quit)
        self._import_worker.finished.connect(self._import_worker.deleteLater)
        self._import_thread.finished.connect(self._import_thread.deleteLater)
        self._progress.canceled.connect(self._import_worker.cancel)
        self._import_thread.start()

    def _on_import_progress(self, done: int, total: int):
        if self._progress is not None:
            self._progress.setValue(done)

    def _on_import_finished(self, result: dict, pre_rejected: list[tuple[str, str]]):
        if self._progress is not None:
            self._progress.close()
            self._progress = None
        self._import_thread = None
        self._import_worker = None
        self._show_import_summary(result, pre_rejected)
        self._do_search()

    def _show_import_summary(self, result: dict, pre_rejected: list[tuple[str, str]]):
        created = int(result.get("created", 0))
        rejected = list(pre_rejected) + list(result.get("rejected", []))
        cancelled = bool(result.get("cancelled"))

        head = f"Criados: {created}\nRejeitados: {len(rejected)}"
        if cancelled:
            head = "Importação cancelada.\n\n" + head

        box = QMessageBox(self)
        box.setWindowTitle("Resultado da importação")
        box.setIcon(QMessageBox.Icon.Information if not rejected else QMessageBox.Icon.Warning)
        box.setText(head)
        if rejected:
            lines = [f"• {code or '(sem código)'} — {reason}" for code, reason in rejected]
            box.setDetailedText("\n".join(lines))
        box.exec()

    # ── Infra de threads ──────────────────────────────────────────────────────
    def _run_action(self, fn, *args, on_result=None, success_message: str = ""):
        worker = ActionWorker(fn, *args)
        thread = QThread()
        cb = UiCallback()
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.result.connect(cb.result)
        worker.error.connect(cb.error)
        if on_result:
            cb.result.connect(on_result)
        if success_message:
            cb.result.connect(
                lambda _: QMessageBox.information(self, "Cadastro de clientes", success_message)
            )
        cb.error.connect(self._show_error)
        cb.error.connect(lambda msg: QMessageBox.critical(self, "Cadastro de clientes", msg))
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda t=thread, w=worker: self._cleanup_thread(t, w))
        worker._cb = cb
        thread.start()
        self._threads.append((thread, worker))

    def _cleanup_thread(self, thread: QThread, worker: QObject):
        self._threads = [pair for pair in self._threads if pair != (thread, worker)]

    def _show_error(self, message: str):
        self.error_label.setText(f"Não foi possível carregar os clientes.\n\n{message}")
        self.error_label.show()

    # ── Tema ──────────────────────────────────────────────────────────────────
    def _apply_table_style(self) -> None:
        s = self.scale
        self.table.setStyleSheet(theme.neon_table_qss(self.scale))
        theme.apply_neon_table_palette(self.table)

    def apply_theme(self) -> None:
        s = self.scale
        bg = theme.CONTENT_BG
        self.setStyleSheet(f"QWidget#clientCenterView {{ background:{bg}; }}")
        self._page_scroll.setStyleSheet(f"QScrollArea {{ border:none; background:{bg}; }}")
        self._page_scroll.viewport().setStyleSheet(f"background:{bg}; border:none;")
        self._page_content.setStyleSheet(f"QWidget#clientCenterContent {{ background:{bg}; }}")
        self.error_label.setStyleSheet(
            f"background:{_rgba(theme.DANGER, 18)}; color:{theme.DANGER};"
            f"border:1px solid {_rgba(theme.DANGER, 48)}; border-radius:16px;"
            f"padding:12px 14px; font-size:{max(8, int(9 * s))}pt; font-weight:600;"
        )
        self.search_input.setStyleSheet(_field_style(s))
        self._apply_table_style()
        self.btn_template.setStyleSheet(_flat_secondary_btn_style(s))
        self.btn_import.setStyleSheet(_flat_secondary_btn_style(s))
        for inp in (self.input_code, self.input_name, self.input_cnpj):
            inp.setStyleSheet(_field_style(s))
        if self.input_code.isReadOnly():
            self.input_code.setStyleSheet(
                _field_style(s) + "QLineEdit { color:" + theme.TEXT_MEDIUM + "; }"
            )
        self.check_active.setStyleSheet(f"background:transparent; font-size:{max(8, int(9 * s))}pt;")
        self.btn_save.setStyleSheet(_primary_action_btn_style(s))
        self.btn_disable.setStyleSheet(_danger_action_btn_style(s))
