from PySide6.QtCore import QDate, QObject, QThread, Qt, QTimer, Signal
from PySide6.QtGui import QColor, QPalette
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDateEdit,
    QDateTimeEdit,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from ..api import client as api
from ..core import theme
from ..core.session import session
from ..widgets.smooth_scroll import apply_smooth_scroll
from ..widgets.sortable_item import SortableItem
from ..core.datetime_utils import (
    format_date as _format_date,
    format_datetime as _format_datetime,
    format_header_date as _format_header_date,
    local_now,
)


COLS = ["PED", "CLIENTE", "OBRA", "VENDEDOR", "DATA", "STATUS", "FATURADO", "PRODUÇÃO", "MÁQUINA", "MOTIVO"]

PRODUCTION_OPTIONS = (
    ("Todas as produções", ""),
    ("A&R", "A&R"),
    ("Pinheiro Indústria", "Pinheiro Indústria"),
)

if "FATURADO" not in COLS:
    COLS.insert(6, "FATURADO")

_LEGACY_INVOICED_OPTIONS = (
    ("TODOS", ""),
    ("SIM", "sim"),
    ("NÃO", "nao"),
)

INVOICED_OPTIONS = (
    ("TODOS", ""),
    ("SIM", "sim"),
    ("NÃO", "nao"),
)

ALL_DATES_SENTINEL = QDate(2000, 1, 1)

STATUS_LABELS = {
    **theme.STATUS_LABELS,
    "finalizada_producao": "Finalizada na Produção",
    "cancelada_producao": "Cancelada na Produção",
}


def _rgba(color: str, alpha: int) -> str:
    parsed = QColor(color)
    return f"rgba({parsed.red()}, {parsed.green()}, {parsed.blue()}, {alpha})"


def _apply_shadow(widget: QWidget, blur: int = 28, y_offset: int = 6, alpha: int = 24) -> None:
    shadow = QGraphicsDropShadowEffect(widget)
    shadow.setBlurRadius(blur)
    shadow.setOffset(0, y_offset)
    color = QColor(theme.TEXT_DARK)
    color.setAlpha(alpha)
    shadow.setColor(color)
    widget.setGraphicsEffect(shadow)


def _make_card(
    scale: float,
    background: str | None = None,
    border_color: str | None = None,
    radius: int = 18,
    hover_background: str | None = None,
) -> QFrame:
    card = QFrame()
    card.setObjectName("historyCard")
    card.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
    card.setAttribute(Qt.WidgetAttribute.WA_Hover, True)
    card.setProperty("theme_bg", "card_bordered" if border_color else "card")
    card.setStyleSheet(f"QFrame#historyCard {{ border-radius:{radius}px; }}")
    _apply_shadow(card, blur=max(26, int(30 * scale)), y_offset=max(4, int(5 * scale)))
    return card


def _flat_secondary_btn_style(scale: float) -> str:
    fs = max(9, int(10 * scale))
    return (
        f"QPushButton {{"
        f"  background:{theme.CARD_BG}; color:{theme.TEXT_DARK};"
        f"  border:1px solid {theme.BORDER_COLOR}; outline:none; border-radius:14px;"
        f"  padding:9px 18px; font-size:{fs}pt; font-weight:700;"
        f"}}"
        f"QPushButton:hover {{ background:{theme.TABLE_ALT_ROW}; border-color:{_rgba(theme.PRIMARY, 70)}; }}"
        f"QPushButton:pressed {{ background:#E7EEF7; }}"
        f"QPushButton:disabled {{ background:#E5EAF2; color:#97A3B6; border-color:#E5EAF2; }}"
    )


def _primary_action_btn_style(scale: float) -> str:
    fs = max(9, int(10 * scale))
    return (
        f"QPushButton {{"
        f"  background:{theme.PRIMARY}; color:#FFFFFF; border:none; border-radius:14px;"
        f"  padding:9px 18px; font-size:{fs}pt; font-weight:700;"
        f"}}"
        f"QPushButton:hover {{ background:{theme.PRIMARY_HOVER}; }}"
        f"QPushButton:pressed {{ background:#152D49; }}"
        f"QPushButton:disabled {{ background:#A7B3C6; color:#F8FAFC; }}"
    )


def _field_style(scale: float) -> str:
    fs = max(9, int(10 * scale))
    arrow = _rgba(theme.TEXT_LIGHT, 220)
    arrow_hover = _rgba(theme.TEXT_DARK, 210)
    btn_bg = _rgba(theme.PRIMARY, 18)
    btn_bg_hover = _rgba(theme.PRIMARY, 28)
    return (
        f"QLineEdit, QComboBox, QDateEdit {{"
        f"  background:{theme.CARD_BG}; border:1px solid {theme.BORDER_COLOR}; border-radius:14px;"
        f"  padding:9px 12px; font-size:{fs}pt; color:{theme.TEXT_DARK};"
        f"  selection-background-color:{_rgba(theme.PRIMARY, 24)}; selection-color:{theme.TEXT_DARK};"
        f"}}"
        f"QComboBox {{ padding-right:34px; }}"
        f"QDateEdit {{ padding-right:34px; }}"
        f"QLineEdit {{ placeholder-text-color:{theme.TEXT_MEDIUM}; }}"
        f"QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{ border:1px solid {_rgba(theme.PRIMARY, 88)}; }}"
        f"QComboBox::drop-down {{"
        f"  subcontrol-origin:padding; subcontrol-position:top right;"
        f"  width:24px; border:none; margin:3px 6px 3px 0;"
        f"  border-radius:8px; background:{btn_bg};"
        f"}}"
        f"QComboBox::drop-down:hover {{ background:{btn_bg_hover}; }}"
        f"QComboBox::down-arrow {{"
        f"  width:0px; height:0px;"
        f"  border-left:5px solid transparent; border-right:5px solid transparent;"
        f"  border-top:6px solid {arrow};"
        f"}}"
        f"QComboBox::down-arrow:on {{ border-top:6px solid {arrow_hover}; }}"
        f"QDateEdit::up-button, QDateEdit::down-button {{"
        f"  subcontrol-origin:border; width:18px;"
        f"  border:none; background:{btn_bg};"
        f"}}"
        f"QDateEdit::up-button {{ subcontrol-position:top right; margin:3px 6px 1px 0; border-top-left-radius:7px; border-top-right-radius:7px; }}"
        f"QDateEdit::down-button {{ subcontrol-position:bottom right; margin:1px 6px 3px 0; border-bottom-left-radius:7px; border-bottom-right-radius:7px; }}"
        f"QDateEdit::up-button:hover, QDateEdit::down-button:hover {{ background:{btn_bg_hover}; }}"
        f"QDateEdit::up-arrow {{"
        f"  width:0px; height:0px;"
        f"  border-left:4px solid transparent; border-right:4px solid transparent;"
        f"  border-bottom:6px solid {arrow};"
        f"}}"
        f"QDateEdit::down-arrow {{"
        f"  width:0px; height:0px;"
        f"  border-left:4px solid transparent; border-right:4px solid transparent;"
        f"  border-top:6px solid {arrow};"
        f"}}"
        f"QComboBox QAbstractItemView, QDateEdit QAbstractItemView {{"
        f"  background:{theme.CARD_BG}; color:{theme.TEXT_DARK}; border:1px solid {theme.BORDER_COLOR};"
        f"  selection-background-color:{_rgba(theme.PRIMARY, 18)}; selection-color:{theme.TEXT_DARK};"
        f"}}"
    )


def _calendar_btn_style(scale: float) -> str:
    fs = max(11, int(13 * scale))
    return (
        f"QToolButton {{"
        f"  background:{theme.PRIMARY}; color:#FFFFFF; border:none; border-radius:12px;"
        f"  font-size:{fs}pt; font-weight:700; padding:0px 2px;"
        f"}}"
        f"QToolButton:hover {{ background:{theme.PRIMARY_HOVER}; }}"
        f"QToolButton:pressed {{ background:#152D49; }}"
    )


class PeriodDateEdit(QDateEdit):
    def mousePressEvent(self, event) -> None:
        super().mousePressEvent(event)
        QTimer.singleShot(0, self._prioritize_day_section)
        QTimer.singleShot(0, self._select_all_text)

    def focusInEvent(self, event) -> None:
        super().focusInEvent(event)
        QTimer.singleShot(0, self._prioritize_day_section)
        QTimer.singleShot(0, self._select_all_text)

    def stepBy(self, steps: int) -> None:
        self._prioritize_day_section()
        super().stepBy(steps)

    def keyPressEvent(self, event) -> None:
        if event.modifiers() in (
            Qt.KeyboardModifier.NoModifier,
            Qt.KeyboardModifier.ShiftModifier,
        ):
            key_text = (event.text() or "").strip().lower()
            if key_text == "h":
                self._set_relative_date(0)
                return
            if key_text == "o":
                self._set_relative_date(-1)
                return
            if key_text == "i":
                self._set_today_anchor(day=1)
                return
            if key_text == "f":
                self._set_end_of_current_month()
                return
            if key_text == "a":
                self._set_today_anchor(month=1, day=1)
                return
        super().keyPressEvent(event)

    def _set_relative_date(self, days: int) -> None:
        today = local_now().date()
        chosen = QDate(today.year, today.month, today.day).addDays(days)
        self.setDate(chosen)
        self._prioritize_day_section()
        self._select_all_text()

    def _set_today_anchor(self, *, month: int | None = None, day: int | None = None) -> None:
        today = local_now().date()
        chosen = QDate(today.year, month or today.month, day or today.day)
        self.setDate(chosen)
        self._prioritize_day_section()
        self._select_all_text()

    def _set_end_of_current_month(self) -> None:
        today = local_now().date()
        chosen = QDate(today.year, today.month, 1)
        chosen = chosen.addMonths(1).addDays(-1)
        self.setDate(chosen)
        self._prioritize_day_section()
        self._select_all_text()

    def _select_all_text(self) -> None:
        editor = self.lineEdit()
        if editor is not None:
            editor.selectAll()

    def _prioritize_day_section(self) -> None:
        self.setCurrentSection(QDateTimeEdit.Section.DaySection)


class HistoryWorker(QObject):
    result = Signal(object)
    error = Signal(str)
    finished = Signal()

    def __init__(
        self,
        status: str = "",
        search: str = "",
        emission_date_start: str = "",
        emission_date_end: str = "",
        production_destination: str = "",
        production_machine: str = "",
        invoiced: str = "",
    ):
        super().__init__()
        self.status = status
        self.search = search
        self.emission_date_start = emission_date_start
        self.emission_date_end = emission_date_end
        self.production_destination = production_destination
        self.production_machine = production_machine
        self.invoiced = invoiced

    def run(self):
        try:
            invoiced_value = None
            if self.invoiced == "sim":
                invoiced_value = True
            elif self.invoiced == "nao":
                invoiced_value = False
            self.result.emit(
                api.list_requisitions(
                    self.status,
                    self.search,
                    limit=100,
                    emission_date_start=self.emission_date_start,
                    emission_date_end=self.emission_date_end,
                    production_destination=self.production_destination,
                    production_machine=self.production_machine,
                    invoiced=invoiced_value,
                )
            )
        except Exception as exc:
            self.error.emit(str(exc))
        finally:
            self.finished.emit()


class MachineOptionsWorker(QObject):
    result = Signal(object)
    error = Signal(str)
    finished = Signal()

    def __init__(self, destination: str = ""):
        super().__init__()
        self.destination = destination

    def run(self):
        try:
            if not self.destination:
                self.result.emit([])
                return
            machines = api.get_production_machines(self.destination)
            self.result.emit(machines if isinstance(machines, list) else [])
        except Exception as exc:
            self.error.emit(str(exc))
        finally:
            self.finished.emit()


class HistoryView(QWidget):
    open_requisition = Signal(int)
    guide_requested  = Signal()          # emitido pelo botão ? de ajuda

    def __init__(self, scale: float = 1.0, parent=None):
        super().__init__(parent)
        self.scale = scale
        self._threads: list[tuple[QThread, QObject]] = []
        self._reqs: list[dict] = []
        self._setup_ui()
        self._reset_machine_filter()

    def _setup_ui(self):
        s = self.scale
        page_bg = theme.CONTENT_BG
        self.setObjectName("historyView")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(
            f"QWidget#historyView {{ background:{page_bg}; }}"
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(max(18, int(24 * s)), max(18, int(24 * s)),
                                max(18, int(24 * s)), max(18, int(24 * s)))
        root.setSpacing(max(14, int(18 * s)))

        header = QHBoxLayout()
        header.setSpacing(max(12, int(16 * s)))
        title_col = QVBoxLayout()
        title_col.setSpacing(max(4, int(5 * s)))

        title = QLabel("Histórico / Busca")
        title.setStyleSheet(
            f"font-size:{max(18, int(24 * s))}pt; font-weight:800;"
        )
        subtitle = QLabel(
            "Consulta operacional de requisições por status, pedido, cliente, obra e vendedor."
        )
        subtitle.setWordWrap(True)
        subtitle.setProperty("muted", "1")
        subtitle.setStyleSheet(
            f"font-size:{max(8, int(10 * s))}pt;"
        )
        title_col.addWidget(title)
        title_col.addWidget(subtitle)
        header.addLayout(title_col, 1)

        right_col = QHBoxLayout()
        right_col.setSpacing(max(10, int(12 * s)))

        info_card = _make_card(
            s,
            theme.CARD_BG,
            border_color=None,
            radius=max(16, int(18 * s)),
            hover_background=theme.CARD_BG,
        )
        info_layout = QVBoxLayout(info_card)
        info_layout.setContentsMargins(max(14, int(16 * s)), max(10, int(12 * s)),
                                       max(14, int(16 * s)), max(10, int(12 * s)))
        info_layout.setSpacing(max(2, int(3 * s)))

        date_hint = QLabel("DATA ATUAL")
        date_hint.setProperty("muted", "1")
        date_hint.setStyleSheet(
            f"font-size:{max(7, int(8 * s))}pt; font-weight:700; background:transparent;"
        )
        self.date_label = QLabel(_format_header_date())
        self.date_label.setStyleSheet(
            f"font-size:{max(13, int(16 * s))}pt; font-weight:800; background:transparent;"
        )
        self.updated_label = QLabel("Pronto para consultar")
        self.updated_label.setProperty("muted", "1")
        self.updated_label.setStyleSheet(
            f"font-size:{max(7, int(8 * s))}pt; background:transparent;"
        )
        self.updated_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        info_layout.addWidget(date_hint)
        info_layout.addWidget(self.date_label)
        info_layout.addWidget(self.updated_label)

        self.refresh_btn = QPushButton("ATUALIZAR")
        self.refresh_btn.setFixedHeight(max(38, int(44 * s)))
        self.refresh_btn.setStyleSheet(_flat_secondary_btn_style(s))
        self.refresh_btn.clicked.connect(self.refresh)
        right_col.addWidget(info_card)
        right_col.addWidget(self.refresh_btn, 0, Qt.AlignmentFlag.AlignTop)
        header.addLayout(right_col)
        root.addLayout(header)

        self.error_label = QLabel("")
        self.error_label.hide()
        self.error_label.setWordWrap(True)
        self.error_label.setStyleSheet(
            f"background:{_rgba(theme.DANGER, 18)}; color:{theme.DANGER};"
            f"border:1px solid {_rgba(theme.DANGER, 48)}; border-radius:16px;"
            f"padding:12px 14px; font-size:{max(8, int(9 * s))}pt; font-weight:600;"
        )
        root.addWidget(self.error_label)

        filter_card = _make_card(
            s,
            theme.CARD_BG,
            border_color=None,
            radius=max(18, int(20 * s)),
            hover_background=theme.CARD_BG,
        )
        filter_layout = QVBoxLayout(filter_card)
        filter_layout.setContentsMargins(max(16, int(20 * s)), max(14, int(18 * s)),
                                         max(16, int(20 * s)), max(14, int(18 * s)))
        filter_layout.setSpacing(max(10, int(12 * s)))

        filter_accent = QFrame()
        filter_accent.setFixedHeight(max(4, int(5 * s)))
        filter_accent.setStyleSheet(
            f"background:qlineargradient(x1:0, y1:0, x2:1, y2:0,"
            f"stop:0 {_rgba(theme.PRIMARY_HOVER, 235)}, stop:0.5 {_rgba(theme.PRIMARY_HOVER, 155)}, stop:1 {_rgba(theme.PRIMARY_HOVER, 235)});"
            f"border:none; border-radius:{max(2, int(3 * s))}px;"
        )
        filter_layout.addWidget(filter_accent)

        filter_title = QLabel("Filtros de Busca")
        filter_title.setStyleSheet(
            f"font-size:{max(10, int(12 * s))}pt; font-weight:800; background:transparent;"
        )
        filter_shortcuts = QLabel(
            "Atalhos do período: h = hoje | o = ontem | i = início do mês | f = final do mês | a = início do ano"
        )
        filter_shortcuts.setWordWrap(True)
        filter_shortcuts.setStyleSheet(
            f"color:{theme.PRIMARY}; font-size:{max(7, int(8 * s))}pt; font-weight:700; background:transparent;"
        )
        filter_subtitle = QLabel(
            "Refine a consulta por status, produção e máquina, ou pesquise por pedido, cliente ou obra."
        )
        filter_subtitle.setWordWrap(True)
        filter_subtitle.setProperty("muted", "1")
        filter_subtitle.setStyleSheet(
            f"font-size:{max(7, int(8 * s))}pt; background:transparent;"
        )
        filter_layout.addWidget(filter_title)
        filter_layout.addWidget(filter_shortcuts)
        filter_layout.addWidget(filter_subtitle)

        status_col = QVBoxLayout()
        status_col.setSpacing(max(6, int(8 * s)))
        status_label = QLabel("STATUS")
        status_label.setProperty("muted", "1")
        status_label.setStyleSheet(
            f"font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )
        self.combo_status = QComboBox()
        self.combo_status.addItem("Todos os status", "")
        for key, label in theme.STATUS_LABELS.items():
            self.combo_status.addItem(label, key)
        self.combo_status.setFixedHeight(max(38, int(44 * s)))
        self.combo_status.setStyleSheet(_field_style(s))
        status_col.addWidget(status_label)
        status_col.addWidget(self.combo_status)

        period_col = QVBoxLayout()
        period_col.setSpacing(max(6, int(8 * s)))
        period_label = QLabel("PERÍODO")
        period_label.setStyleSheet(
            f"color:{theme.TEXT_MEDIUM}; font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )
        period_row = QHBoxLayout()
        period_row.setSpacing(max(6, int(8 * s)))
        today = local_now().date()
        today_qdate = QDate(today.year, today.month, today.day)

        self.input_date_from = PeriodDateEdit(ALL_DATES_SENTINEL)
        self.input_date_from.setMinimumDate(ALL_DATES_SENTINEL)
        self.input_date_from.setDate(today_qdate)
        self.input_date_from.setSpecialValueText("Data inicial")
        self.input_date_from.setDisplayFormat("dd/MM/yyyy")
        self.input_date_from.setCalendarPopup(False)
        self.input_date_from.setFixedHeight(max(38, int(44 * s)))
        self.input_date_from.setStyleSheet(_field_style(s))
        self.btn_date_from = QToolButton()
        self.btn_date_from.setText("\U0001F4C5")
        self.btn_date_from.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_date_from.setToolTip("Usar a data de hoje")
        self.btn_date_from.setFixedSize(max(38, int(42 * s)), max(38, int(44 * s)))
        self.btn_date_from.setStyleSheet(_calendar_btn_style(s))
        self.btn_date_from.clicked.connect(lambda: self._set_date_today(self.input_date_from))

        until_label = QLabel("ATÉ")
        until_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        until_label.setStyleSheet(
            f"color:{theme.TEXT_MEDIUM}; font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )

        self.input_date_to = PeriodDateEdit(ALL_DATES_SENTINEL)
        self.input_date_to.setMinimumDate(ALL_DATES_SENTINEL)
        self.input_date_to.setDate(today_qdate)
        self.input_date_to.setSpecialValueText("Data final")
        self.input_date_to.setDisplayFormat("dd/MM/yyyy")
        self.input_date_to.setCalendarPopup(False)
        self.input_date_to.setFixedHeight(max(38, int(44 * s)))
        self.input_date_to.setStyleSheet(_field_style(s))
        self.btn_date_to = QToolButton()
        self.btn_date_to.setText("\U0001F4C5")
        self.btn_date_to.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_date_to.setToolTip("Usar a data de hoje")
        self.btn_date_to.setFixedSize(max(38, int(42 * s)), max(38, int(44 * s)))
        self.btn_date_to.setStyleSheet(_calendar_btn_style(s))
        self.btn_date_to.clicked.connect(lambda: self._set_date_today(self.input_date_to))

        period_row.addWidget(self.input_date_from, 1)
        period_row.addWidget(self.btn_date_from)
        period_row.addWidget(until_label)
        period_row.addWidget(self.input_date_to, 1)
        period_row.addWidget(self.btn_date_to)
        period_col.addWidget(period_label)
        period_col.addLayout(period_row)

        invoiced_col = QVBoxLayout()
        invoiced_col.setSpacing(max(6, int(8 * s)))
        invoiced_label = QLabel("FATURADO")
        invoiced_label.setStyleSheet(
            f"color:{theme.TEXT_MEDIUM}; font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )
        self.combo_invoiced = QComboBox()
        for label, value in INVOICED_OPTIONS:
            self.combo_invoiced.addItem(label, value)
        self.combo_invoiced.setFixedHeight(max(38, int(44 * s)))
        self.combo_invoiced.setStyleSheet(_field_style(s))
        invoiced_col.addWidget(invoiced_label)
        invoiced_col.addWidget(self.combo_invoiced)

        production_col = QVBoxLayout()
        production_col.setSpacing(max(6, int(8 * s)))
        production_label = QLabel("PRODUÇÃO")
        production_label.setStyleSheet(
            f"color:{theme.TEXT_MEDIUM}; font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )
        self.combo_production = QComboBox()
        for label, value in PRODUCTION_OPTIONS:
            self.combo_production.addItem(label, value)
        self.combo_production.setFixedHeight(max(38, int(44 * s)))
        self.combo_production.setStyleSheet(_field_style(s))
        self.combo_production.currentIndexChanged.connect(self._on_production_changed)
        production_col.addWidget(production_label)
        production_col.addWidget(self.combo_production)

        machine_col = QVBoxLayout()
        machine_col.setSpacing(max(6, int(8 * s)))
        machine_label = QLabel("MÁQUINA")
        machine_label.setStyleSheet(
            f"color:{theme.TEXT_MEDIUM}; font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )
        self.combo_machine = QComboBox()
        self.combo_machine.setFixedHeight(max(38, int(44 * s)))
        self.combo_machine.setStyleSheet(_field_style(s))
        machine_col.addWidget(machine_label)
        machine_col.addWidget(self.combo_machine)

        search_col = QVBoxLayout()
        search_col.setSpacing(max(6, int(8 * s)))
        search_label = QLabel("BUSCA")
        search_label.setProperty("muted", "1")
        search_label.setStyleSheet(
            f"font-size:{max(7, int(8 * s))}pt; font-weight:700;"
        )
        self.input_search = QLineEdit()
        self.input_search.setPlaceholderText("Buscar por PED, cliente ou obra...")
        self.input_search.setFixedHeight(max(38, int(44 * s)))
        self.input_search.setStyleSheet(_field_style(s))
        self.input_search.returnPressed.connect(self.refresh)
        search_col.addWidget(search_label)
        search_col.addWidget(self.input_search)

        buttons_col = QVBoxLayout()
        buttons_col.setSpacing(max(6, int(8 * s)))
        buttons_col.addWidget(QLabel(""))
        buttons_row = QHBoxLayout()
        buttons_row.setSpacing(max(8, int(10 * s)))
        self.search_btn = QPushButton("BUSCAR")
        self.search_btn.setFixedHeight(max(38, int(44 * s)))
        self.search_btn.setStyleSheet(_primary_action_btn_style(s))
        self.search_btn.clicked.connect(self.refresh)

        self.clear_btn = QPushButton("LIMPAR")
        self.clear_btn.setFixedHeight(max(38, int(44 * s)))
        self.clear_btn.setStyleSheet(_flat_secondary_btn_style(s))
        self.clear_btn.clicked.connect(self._clear_filters)

        buttons_row.addWidget(self.search_btn)
        buttons_row.addWidget(self.clear_btn)
        buttons_col.addLayout(buttons_row)

        # Botão ? — abre o guia rápido desta tela
        sz_g = max(24, int(28 * s))
        self.guide_btn = QPushButton("?")
        self.guide_btn.setToolTip("Abrir guia rápido")
        self.guide_btn.setFixedSize(sz_g, sz_g)
        self.guide_btn.setStyleSheet(
            f"font-size:{max(10, int(11 * s))}pt; font-weight:700;"
            f"color:{theme.TEXT_MEDIUM}; background:transparent;"
            f"border:1px solid {theme.BORDER_COLOR};"
            f"border-radius:{sz_g // 2}px; padding:0;"
        )
        self.guide_btn.clicked.connect(self.guide_requested)
        buttons_col.addWidget(self.guide_btn, 0, Qt.AlignmentFlag.AlignLeft)

        # Filtros em duas linhas, agrupadas por intenção, para dar respiro e
        # acomodar novos filtros no futuro sem espremer tudo numa linha só.
        # Linha 1 — o que/quando buscar: BUSCA · PERÍODO · STATUS
        row1 = QHBoxLayout()
        row1.setSpacing(max(12, int(16 * s)))
        row1.addLayout(search_col, 2)
        row1.addLayout(period_col, 2)
        row1.addLayout(status_col, 1)

        # Linha 2 — refinamentos + ações: FATURADO · PRODUÇÃO · MÁQUINA · [BUSCAR/LIMPAR]
        row2 = QHBoxLayout()
        row2.setSpacing(max(12, int(16 * s)))
        row2.addLayout(invoiced_col, 1)
        row2.addLayout(production_col, 1)
        row2.addLayout(machine_col, 1)
        row2.addStretch(1)
        row2.addLayout(buttons_col)

        filter_layout.addLayout(row1)
        filter_layout.addLayout(row2)
        root.addWidget(filter_card)

        results_card = _make_card(
            s,
            theme.CARD_BG,
            border_color=None,
            radius=max(18, int(20 * s)),
            hover_background=theme.CARD_BG,
        )
        results_layout = QVBoxLayout(results_card)
        results_layout.setContentsMargins(max(16, int(20 * s)), max(14, int(18 * s)),
                                          max(16, int(20 * s)), max(14, int(18 * s)))
        results_layout.setSpacing(max(10, int(12 * s)))

        results_accent = QFrame()
        results_accent.setFixedHeight(max(4, int(5 * s)))
        results_accent.setStyleSheet(
            f"background:qlineargradient(x1:0, y1:0, x2:1, y2:0,"
            f"stop:0 {_rgba(theme.PRIMARY, 235)}, stop:0.5 {_rgba(theme.PRIMARY, 155)}, stop:1 {_rgba(theme.PRIMARY, 235)});"
            f"border:none; border-radius:{max(2, int(3 * s))}px;"
        )
        results_layout.addWidget(results_accent)

        results_header = QHBoxLayout()
        results_header.setSpacing(max(8, int(10 * s)))
        results_title_col = QVBoxLayout()
        results_title_col.setSpacing(max(2, int(3 * s)))
        results_title = QLabel("Resultados")
        results_title.setStyleSheet(
            f"font-size:{max(10, int(12 * s))}pt; font-weight:800; background:transparent;"
        )
        results_subtitle = QLabel("Duplo clique para abrir a requisição selecionada.")
        results_subtitle.setWordWrap(True)
        results_subtitle.setProperty("muted", "1")
        results_subtitle.setStyleSheet(
            f"font-size:{max(7, int(8 * s))}pt; background:transparent;"
        )
        results_title_col.addWidget(results_title)
        results_title_col.addWidget(results_subtitle)
        results_header.addLayout(results_title_col, 1)

        self.export_btn = QPushButton("EXPORTAR EXCEL")
        self.export_btn.setFixedHeight(max(36, int(42 * s)))
        self.export_btn.setStyleSheet(_flat_secondary_btn_style(s))
        self.export_btn.setToolTip("Exportar os resultados atuais para uma planilha Excel")
        self.export_btn.clicked.connect(self._export_excel)
        results_header.addWidget(self.export_btn, 0, Qt.AlignmentFlag.AlignVCenter)
        results_layout.addLayout(results_header)

        self.table = QTableWidget(0, len(COLS))
        self.table.setHorizontalHeaderLabels(COLS)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setFrameShape(QFrame.Shape.NoFrame)
        self.table.setShowGrid(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        apply_smooth_scroll(self.table)
        header_widget = self.table.horizontalHeader()
        stretch_columns = {1, 2}
        for col in range(len(COLS)):
            mode = (
                QHeaderView.ResizeMode.Stretch
                if col in stretch_columns
                else QHeaderView.ResizeMode.ResizeToContents
            )
            header_widget.setSectionResizeMode(col, mode)
        header_widget.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        header_widget.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        header_widget.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        header_widget.setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)
        header_widget.setSectionResizeMode(7, QHeaderView.ResizeMode.Interactive)
        header_widget.setSectionResizeMode(8, QHeaderView.ResizeMode.Interactive)
        header_widget.setSectionResizeMode(9, QHeaderView.ResizeMode.Interactive)
        header_widget.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header_widget.setMinimumHeight(max(34, int(40 * s)))
        self.table.verticalHeader().setDefaultSectionSize(max(32, int(38 * s)))
        self.table.setColumnWidth(1, max(180, int(220 * s)))
        self.table.setColumnWidth(3, max(150, int(180 * s)))
        self.table.setColumnWidth(5, max(150, int(180 * s)))
        self.table.setColumnWidth(6, max(110, int(130 * s)))
        self.table.setColumnWidth(7, max(150, int(170 * s)))
        self.table.setColumnWidth(8, max(220, int(260 * s)))
        self.table.setColumnWidth(9, max(160, int(200 * s)))
        self.table.setSortingEnabled(True)
        self.table.setStyleSheet(
            f"QTableWidget {{"
            f"  border:none; outline:none; background:{theme.CARD_BG};"
            f"  alternate-background-color:{theme.TABLE_ALT_ROW}; color:{theme.TEXT_DARK};"
            f"  border-radius:14px; gridline-color:transparent; font-size:{max(8, int(9 * s))}pt;"
            f"}}"
            f"QHeaderView::section {{"
            f"  background:{theme.PRIMARY}; color:#fff; padding:9px 10px;"
            f"  font-weight:800; font-size:{max(7, int(8 * s))}pt; border:none;"
            f"}}"
            f"QHeaderView::section:hover {{ background:{theme.PRIMARY_HOVER}; }}"
            f"QTableWidget::item {{"
            f"  background:{theme.CARD_BG}; color:{theme.TEXT_DARK};"
            f"  padding:7px 6px; border-bottom:1px solid {_rgba(theme.PRIMARY, 18)};"
            f"}}"
            f"QTableWidget::item:alternate {{ background:{theme.TABLE_ALT_ROW}; color:{theme.TEXT_DARK}; }}"
            f"QTableWidget::item:selected {{ background:{_rgba(theme.PRIMARY, 18)}; color:{theme.TEXT_DARK}; }}"
        )
        pal = self.table.palette()
        pal.setColor(QPalette.ColorRole.Base, QColor(theme.CARD_BG))
        pal.setColor(QPalette.ColorRole.AlternateBase, QColor(theme.TABLE_ALT_ROW))
        pal.setColor(QPalette.ColorRole.Text, QColor(theme.TEXT_DARK))
        pal.setColor(QPalette.ColorRole.HighlightedText, QColor(theme.TEXT_DARK))
        pal.setColor(QPalette.ColorRole.Highlight, QColor(_rgba(theme.PRIMARY, 40)))
        self.table.setPalette(pal)
        self.table.viewport().setAutoFillBackground(True)
        self.table.setMinimumHeight(max(300, int(360 * s)))
        self.table.doubleClicked.connect(self._on_double_click)
        results_layout.addWidget(self.table, 1)
        root.addWidget(results_card, 1)

    def refresh(self):
        self.error_label.hide()
        period = self._selected_emission_period()
        if period is None:
            self.error_label.setText("A data inicial não pode ser maior que a data final.")
            self.error_label.show()
            return
        self._set_loading(True)
        status = self.combo_status.currentData() or ""
        emission_date_start, emission_date_end = period
        invoiced = self.combo_invoiced.currentData() or ""
        production_destination = self.combo_production.currentData() or ""
        production_machine = self.combo_machine.currentData() or ""
        search = self.input_search.text().strip()

        worker = HistoryWorker(
            status,
            search,
            emission_date_start,
            emission_date_end,
            production_destination,
            production_machine,
            invoiced,
        )
        thread = QThread()
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.result.connect(self._populate)
        worker.error.connect(self._show_error)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda t=thread, w=worker: self._cleanup_thread(t, w))
        thread.finished.connect(lambda: self._set_loading(False))
        thread.start()
        self._threads.append((thread, worker))

    def _cleanup_thread(self, thread: QThread, worker: QObject):
        self._threads = [pair for pair in self._threads if pair != (thread, worker)]

    def _selected_emission_period(self) -> tuple[str, str] | None:
        start = self.input_date_from.date()
        end = self.input_date_to.date()
        if start > end:
            return None
        start_value = start.toString("yyyy-MM-dd")
        end_value = end.toString("yyyy-MM-dd")
        return start_value, end_value

    def _focus_date_field(self, field: QDateEdit) -> None:
        field.setFocus(Qt.FocusReason.MouseFocusReason)
        if isinstance(field, PeriodDateEdit):
            field._prioritize_day_section()
        field.selectAll()

    def _set_date_today(self, field: QDateEdit) -> None:
        today = local_now().date()
        field.setDate(QDate(today.year, today.month, today.day))
        self._focus_date_field(field)

    def _reset_machine_filter(self, placeholder: str = "Todas as máquinas"):
        self.combo_machine.blockSignals(True)
        self.combo_machine.clear()
        self.combo_machine.addItem(placeholder, "")
        self.combo_machine.setCurrentIndex(0)
        self.combo_machine.blockSignals(False)
        self.combo_machine.setEnabled(False)

    def _on_production_changed(self):
        destination = self.combo_production.currentData() or ""
        if not destination:
            self._reset_machine_filter()
            return
        self._load_machine_options(str(destination))

    def _load_machine_options(self, destination: str):
        self.error_label.hide()
        self.combo_machine.blockSignals(True)
        self.combo_machine.clear()
        self.combo_machine.addItem("Carregando máquinas...", "")
        self.combo_machine.setCurrentIndex(0)
        self.combo_machine.blockSignals(False)
        self.combo_machine.setEnabled(False)

        worker = MachineOptionsWorker(destination)
        thread = QThread()
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.result.connect(
            lambda machines, selected=destination: self._populate_machine_options(selected, machines)
        )
        worker.error.connect(
            lambda message, selected=destination: self._handle_machine_options_error(selected, message)
        )
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda t=thread, w=worker: self._cleanup_thread(t, w))
        thread.start()
        self._threads.append((thread, worker))

    def _populate_machine_options(self, destination: str, machines: object):
        current_destination = str(self.combo_production.currentData() or "")
        if current_destination != destination:
            return

        self.combo_machine.blockSignals(True)
        self.combo_machine.clear()
        self.combo_machine.addItem("Todas as máquinas", "")
        for machine_name in machines if isinstance(machines, list) else []:
            if not isinstance(machine_name, str):
                continue
            name = machine_name.strip()
            if name:
                self.combo_machine.addItem(name, name)
        self.combo_machine.setCurrentIndex(0)
        self.combo_machine.blockSignals(False)
        self.combo_machine.setEnabled(True)

    def _handle_machine_options_error(self, destination: str, message: str):
        current_destination = str(self.combo_production.currentData() or "")
        if current_destination != destination:
            return
        self._reset_machine_filter("Todas as máquinas")
        self.error_label.setText(
            f"Não foi possível carregar as máquinas da produção selecionada.\n\n{message}"
        )
        self.error_label.show()

    def _set_loading(self, loading: bool):
        self.refresh_btn.setEnabled(not loading)
        self.search_btn.setEnabled(not loading)
        self.clear_btn.setEnabled(not loading)
        self.export_btn.setEnabled(not loading)
        self.combo_status.setEnabled(not loading)
        self.input_date_from.setEnabled(not loading)
        self.input_date_to.setEnabled(not loading)
        self.btn_date_from.setEnabled(not loading)
        self.btn_date_to.setEnabled(not loading)
        self.combo_invoiced.setEnabled(not loading)
        self.combo_production.setEnabled(not loading)
        if not loading:
            self.combo_machine.setEnabled(bool(self.combo_production.currentData()))
        elif self.combo_machine.isEnabled():
            self.combo_machine.setEnabled(False)
        if loading:
            self.updated_label.setText("Atualizando dados...")
            self.date_label.setText(_format_header_date())

    def _show_error(self, message: str):
        self.updated_label.setText("Falha ao atualizar")
        self.error_label.setText(f"Não foi possível carregar o histórico.\n\n{message}")
        self.error_label.show()

    def _row_values(self, req: dict) -> list[str]:
        """Valores de exibição de uma linha (usado na tabela e na exportação)."""
        status = str(req.get("production_status") or req.get("status") or "")
        return [
            str(req.get("ped_number") or "-"),
            str(req.get("client_name") or req.get("client_id") or "-"),
            str(req.get("obra") or "-"),
            str(req.get("vendor_name") or req.get("vendor_id") or "-"),
            _format_date(req.get("emission_date")),
            STATUS_LABELS.get(status, status or "-"),
            "SIM" if bool(req.get("invoiced")) or status == "faturado" else "NÃO",
            str(
                req.get("production_destination_display")
                or req.get("production_destination")
                or "-"
            ),
            str(
                req.get("production_machine_display")
                or req.get("production_machine")
                or "-"
            ),
            str(req.get("cancel_reason") or "-"),
        ]

    def _export_excel(self):
        """Exporta os resultados atualmente carregados para uma planilha .xlsx."""
        if not self._reqs:
            QMessageBox.information(
                self, "Exportar para Excel",
                "Não há resultados para exportar. Faça uma busca primeiro.",
            )
            return
        default_name = f"historico_requisicoes_{local_now().strftime('%Y%m%d_%H%M')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar para Excel", default_name, "Planilha Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Histórico"
            ws.append(COLS)
            header_fill = PatternFill("solid", fgColor="1E3A5F")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for req in self._reqs:
                ws.append(self._row_values(req))
            widths = [10, 34, 26, 22, 12, 24, 10, 18, 24, 26]
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.freeze_panes = "A2"
            wb.save(path)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self, "Exportar para Excel",
                f"Não foi possível salvar a planilha.\n\n{exc}",
            )
            return
        QMessageBox.information(
            self, "Exportar para Excel",
            f"{len(self._reqs)} requisição(ões) exportada(s) com sucesso.",
        )

    def _populate(self, reqs: object):
        if not isinstance(reqs, list):
            self._show_error("Resposta inválida do servidor.")
            self._set_empty_message("Nenhuma requisição encontrada.")
            return

        self._reqs = [req for req in reqs if isinstance(req, dict)]
        self.table.setSortingEnabled(False)
        self.table.clearSpans()
        self.table.setRowCount(0)

        if not self._reqs:
            self._set_empty_message("Nenhuma requisição encontrada.")
        else:
            for req in self._reqs:
                row = self.table.rowCount()
                self.table.insertRow(row)
                status      = str(req.get("production_status") or req.get("status") or "")
                ped_raw     = req.get("ped_number")
                date_raw    = str(req.get("emission_date") or "")
                values = self._row_values(req)
                for col, value in enumerate(values):
                    if col == 5:
                        badge = QLabel(STATUS_LABELS.get(status, status or "-"))
                        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        color_map = {
                            "em_andamento": theme.PRIMARY_HOVER,
                            "aguardando_recebimento": theme.WARNING,
                            "aguardando_na_fila": theme.STATUS_COLORS.get("aguardando_na_fila", theme.WARNING),
                            "aguardando_faturamento": theme.STATUS_COLORS.get("aguardando_faturamento", theme.WARNING),
                            "em_producao": theme.PRIMARY,
                            "faturado": theme.STATUS_COLORS.get("faturado", theme.SUCCESS),
                            "finalizada_producao": theme.SUCCESS,
                            "cancelada_producao": theme.DANGER,
                            "cancelada": theme.DANGER,
                        }
                        color = color_map.get(status, theme.BORDER_COLOR)
                        badge.setStyleSheet(
                            f"background:{_rgba(color, 30)}; color:{color}; border-radius:999px;"
                            f"font-weight:700; padding:4px 10px; font-size:{max(7, int(8 * self.scale))}pt;"
                        )
                        self.table.setCellWidget(row, col, badge)
                    elif col == 0:
                        # Ordena PED numericamente
                        try:
                            sort_key = int(ped_raw)
                        except (TypeError, ValueError):
                            sort_key = 0
                        item = SortableItem(value, sort_key=sort_key)
                        req_id = req.get("id")
                        if req_id is not None:
                            item.setData(Qt.ItemDataRole.UserRole, int(req_id))
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.table.setItem(row, col, item)
                    elif col == 4:
                        # Ordena DATA pela string ISO (YYYY-MM-DD... ordena lexicograficamente)
                        item = SortableItem(value, sort_key=date_raw)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.table.setItem(row, col, item)
                    else:
                        item = QTableWidgetItem(value)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.table.setItem(row, col, item)
        self.table.setSortingEnabled(True)

        current = local_now()
        self.date_label.setText(_format_header_date(current))
        self.updated_label.setText(f"Atualizado em {_format_datetime(current)}")

    def _set_empty_message(self, message: str):
        self.table.setRowCount(1)
        self.table.setSpan(0, 0, 1, len(COLS))
        item = QTableWidgetItem(message)
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        item.setFlags(Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(0, 0, item)

    def _on_double_click(self, index):
        row = index.row()
        if row < 0:
            return
        ped_item = self.table.item(row, 0)
        if ped_item is not None:
            req_id = ped_item.data(Qt.ItemDataRole.UserRole)
            if req_id is not None:
                self.open_requisition.emit(int(req_id))
                return
        if 0 <= row < len(self._reqs):
            req_id = self._reqs[row].get("id")
            if req_id is not None:
                self.open_requisition.emit(int(req_id))

    def _clear_filters(self):
        self.combo_status.setCurrentIndex(0)
        today = local_now().date()
        current = QDate(today.year, today.month, today.day)
        self.input_date_from.setDate(current)
        self.input_date_to.setDate(current)
        self.combo_invoiced.setCurrentIndex(0)
        self.combo_production.setCurrentIndex(0)
        self._reset_machine_filter()
        self.input_search.clear()
        self.refresh()

    def _apply_table_style(self) -> None:
        s = self.scale
        self.table.setStyleSheet(
            f"QTableWidget {{"
            f"  border:none; outline:none; background:{theme.CARD_BG};"
            f"  alternate-background-color:{theme.TABLE_ALT_ROW}; color:{theme.TEXT_DARK};"
            f"  border-radius:14px; gridline-color:transparent; font-size:{max(8, int(9 * s))}pt;"
            f"}}"
            f"QHeaderView::section {{"
            f"  background:{theme.PRIMARY}; color:#fff; padding:9px 10px;"
            f"  font-weight:800; font-size:{max(7, int(8 * s))}pt; border:none;"
            f"}}"
            f"QTableWidget::item {{"
            f"  background:{theme.CARD_BG}; color:{theme.TEXT_DARK};"
            f"  padding:7px 6px; border-bottom:1px solid {_rgba(theme.PRIMARY, 18)};"
            f"}}"
            f"QTableWidget::item:alternate {{ background:{theme.TABLE_ALT_ROW}; color:{theme.TEXT_DARK}; }}"
            f"QTableWidget::item:selected {{ background:{_rgba(theme.PRIMARY, 18)}; color:{theme.TEXT_DARK}; }}"
        )
        pal = self.table.palette()
        pal.setColor(QPalette.ColorRole.Base, QColor(theme.CARD_BG))
        pal.setColor(QPalette.ColorRole.AlternateBase, QColor(theme.TABLE_ALT_ROW))
        pal.setColor(QPalette.ColorRole.Text, QColor(theme.TEXT_DARK))
        pal.setColor(QPalette.ColorRole.HighlightedText, QColor(theme.TEXT_DARK))
        pal.setColor(QPalette.ColorRole.Highlight, QColor(_rgba(theme.PRIMARY, 40)))
        self.table.setPalette(pal)

    def apply_theme(self) -> None:
        s = self.scale
        bg = theme.CONTENT_BG
        self.setStyleSheet(f"QWidget#historyView {{ background:{bg}; }}")
        self.refresh_btn.setStyleSheet(_flat_secondary_btn_style(s))
        self.error_label.setStyleSheet(
            f"background:{_rgba(theme.DANGER, 18)}; color:{theme.DANGER};"
            f"border:1px solid {_rgba(theme.DANGER, 48)}; border-radius:16px;"
            f"padding:12px 14px; font-size:{max(8, int(9 * s))}pt; font-weight:600;"
        )
        self.combo_status.setStyleSheet(_field_style(s))
        self.input_date_from.setStyleSheet(_field_style(s))
        self.input_date_to.setStyleSheet(_field_style(s))
        self.btn_date_from.setStyleSheet(_calendar_btn_style(s))
        self.btn_date_to.setStyleSheet(_calendar_btn_style(s))
        self.combo_invoiced.setStyleSheet(_field_style(s))
        self.combo_production.setStyleSheet(_field_style(s))
        self.combo_machine.setStyleSheet(_field_style(s))
        self.input_search.setStyleSheet(_field_style(s))
        self.search_btn.setStyleSheet(_primary_action_btn_style(s))
        self.clear_btn.setStyleSheet(_flat_secondary_btn_style(s))
        self.export_btn.setStyleSheet(_flat_secondary_btn_style(s))
        self._apply_table_style()
